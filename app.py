import os
import sqlite3
from datetime import datetime, timedelta, date
from functools import wraps

from flask import (
    Flask, g, redirect, render_template, request, session,
    url_for, flash, jsonify, send_file
)
from authlib.integrations.flask_client import OAuth
import io
from openpyxl import Workbook, load_workbook

# =========================
# Config
# =========================
APP_SECRET = os.environ.get("APP_SECRET", "crm_mimovil_clave_larga_cambiar")
DATABASE = os.environ.get("DATABASE_PATH", "crm.sqlite")

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")

ALLOWED_EMAILS = set(
    e.strip().lower()
    for e in os.environ.get("ALLOWED_EMAILS", "").split(",")
    if e.strip()
)

ALERT_DAYS = int(os.environ.get("ALERT_DAYS", "30"))

app = Flask(__name__)
app.secret_key = APP_SECRET

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)

# =========================
# OAuth Google
# =========================
oauth = OAuth(app)

google = oauth.register(
    name="google",
    client_id=GOOGLE_CLIENT_ID,
    client_secret=GOOGLE_CLIENT_SECRET,
    access_token_url="https://oauth2.googleapis.com/token",
    authorize_url="https://accounts.google.com/o/oauth2/v2/auth",
    api_base_url="https://www.googleapis.com/oauth2/v2/",
    client_kwargs={"scope": "email profile"},
)

# =========================
# DB helpers
# =========================
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON;")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def _col_exists(db, table: str, col: str) -> bool:
    rows = db.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r["name"] == col for r in rows)


def _add_col_if_missing(db, table: str, col: str, coltype: str):
    if not _col_exists(db, table, col):
        db.execute(f"ALTER TABLE {table} ADD COLUMN {col} {coltype};")


def init_db():
    db = get_db()

    # ---- Clients ----
    db.execute("""
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            dni TEXT NOT NULL,
            birth_date TEXT,
            phone TEXT,
            address TEXT,
            email TEXT,
            current_operator TEXT,
            current_tariff_price TEXT,
            permanence TEXT,
            terminal TEXT,
            sales_done TEXT,
            repairs_done TEXT,
            procedures_done TEXT,
            observations TEXT,
            pending_tasks TEXT,
            created_at TEXT NOT NULL
        );
    """)

    # Compat columnas viejas
    _add_col_if_missing(db, "clients", "permanence_start", "TEXT")
    _add_col_if_missing(db, "clients", "permanence_end", "TEXT")

    # Columnas nuevas permanencia
    _add_col_if_missing(db, "clients", "permanence_start_date", "TEXT")
    _add_col_if_missing(db, "clients", "permanence_months", "INTEGER")
    _add_col_if_missing(db, "clients", "permanence_end_date", "TEXT")

    # Comercial + Estado
    _add_col_if_missing(db, "clients", "commercial", "TEXT")
    _add_col_if_missing(db, "clients", "status", "TEXT")
    _add_col_if_missing(db, "clients", "service_type", "TEXT")

    # ---- Mobile lines ----
    db.execute("""
        CREATE TABLE IF NOT EXISTS mobile_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            line_number TEXT,
            pin TEXT,
            puk TEXT,
            icc TEXT,
            google_or_iphone_account TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE
        );
    """)
    _add_col_if_missing(db, "mobile_lines", "permanence_end_date", "TEXT")

    # ---- Repairs ----
    db.execute("""
        CREATE TABLE IF NOT EXISTS repairs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            date TEXT,
            model TEXT,
            repair TEXT,
            cost REAL,
            created_at TEXT NOT NULL,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE
        );
    """)

    # ---- Sales ----
    db.execute("""
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            date TEXT,
            item TEXT,
            operator TEXT,
            amount REAL,
            notes TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE
        );
    """)

    # Backfill: sincronizar fin permanencia vieja/nueva
    db.execute("""
        UPDATE clients
        SET permanence_end_date = permanence_end
        WHERE (permanence_end_date IS NULL OR permanence_end_date = '')
          AND permanence_end IS NOT NULL AND permanence_end != '';
    """)

    db.execute("""
        UPDATE clients
        SET permanence_end = permanence_end_date
        WHERE (permanence_end IS NULL OR permanence_end = '')
          AND permanence_end_date IS NOT NULL AND permanence_end_date != '';
    """)

    db.commit()


with app.app_context():
    init_db()

# =========================
# Date utils
# =========================
def parse_yyyy_mm_dd(s: str):
    s = (s or "").strip()
    if not s:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def add_months(d: date, months: int) -> date:
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    day = d.day

    if m == 12:
        next_month = date(y + 1, 1, 1)
    else:
        next_month = date(y, m + 1, 1)

    last_day = (next_month - timedelta(days=1)).day
    if day > last_day:
        day = last_day
    return date(y, m, day)


def compute_permanence_end(start_str: str, months_str: str, end_str: str):
    start = parse_yyyy_mm_dd(start_str)
    end = parse_yyyy_mm_dd(end_str)

    months_int = None
    ms = (months_str or "").strip()
    if ms:
        try:
            months_int = int(ms)
        except ValueError:
            months_int = None

    if end is None and start is not None and months_int is not None:
        end = add_months(start, months_int)

    start_iso = start.isoformat() if start else None
    end_iso = end.isoformat() if end else None
    return start_iso, months_int, end_iso


def get_end_date_from_client_row(c):
    if not c:
        return None

    keys = c.keys()
    end_iso = None

    if "permanence_end_date" in keys:
        end_iso = c["permanence_end_date"]

    if (not end_iso) and ("permanence_end" in keys):
        end_iso = c["permanence_end"]

    return (end_iso or "").strip() or None


def days_until(end_iso: str):
    d = parse_yyyy_mm_dd(end_iso)
    if not d:
        return None
    return (d - date.today()).days

# =========================
# Auth
# =========================
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def is_allowed(email):
    if not ALLOWED_EMAILS:
        return True
    return (email or "").lower() in ALLOWED_EMAILS

# =========================
# Routes
# =========================
@app.route("/")
def home():
    if session.get("user"):
        return redirect(url_for("clients"))
    return redirect(url_for("login"))


@app.route("/login")
def login():
    return render_template("login.html")


@app.route("/auth/google")
def auth_google():
    remember = request.args.get("remember") == "1"
    session["remember_me"] = remember
    redirect_uri = url_for("auth_google_callback", _external=True)
    return google.authorize_redirect(redirect_uri)


@app.route("/auth/google/callback")
def auth_google_callback():
    google.authorize_access_token()
    userinfo = google.get("userinfo").json()

    email = userinfo.get("email")
    if not is_allowed(email):
        session.clear()
        flash("Cuenta no autorizada", "danger")
        return redirect(url_for("login"))

    session["user"] = {"email": email, "name": userinfo.get("name") or email}

    if session.get("remember_me"):
        session.permanent = True
        app.permanent_session_lifetime = timedelta(days=30)

    session.pop("remember_me", None)
    return redirect(url_for("clients"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/clients")
@login_required
def clients():
    db = get_db()
    q = request.args.get("q", "").strip()
    only_pending = request.args.get("pending", "0").strip() == "1"
    status_filter = request.args.get("status", "").strip()
    operator_filter = request.args.get("operator", "").strip()
    service_filter = request.args.get("service_type", "").strip()
    start_from = request.args.get("start_from", "").strip()
    start_to = request.args.get("start_to", "").strip()

    where = []
    params = []

    if q:
        where.append("(full_name LIKE ? OR dni LIKE ? OR phone LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

    if only_pending:
        where.append("(pending_tasks IS NOT NULL AND TRIM(pending_tasks) != '')")

    if status_filter:
        where.append("status = ?")
        params.append(status_filter)
    if operator_filter:
        where.append("current_operator = ?")
        params.append(operator_filter)
    if service_filter:
        where.append("service_type = ?")
        params.append(service_filter)

    if start_from:
       where.append("permanence_start_date >= ?")
       params.append(start_from)

    if start_to:
       where.append("permanence_start_date <= ?")
       params.append(start_to)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    rows = db.execute(
        f"SELECT * FROM clients {where_sql} ORDER BY id DESC",
        tuple(params)
    ).fetchall()

    days_left_map = {}
    for c in rows:
        end_iso = get_end_date_from_client_row(c)
        days_left_map[c["id"]] = days_until(end_iso) if end_iso else None

    return render_template(
        "clients_list.html",
        clients=rows,
        q=q,
        alert_days=ALERT_DAYS,
        days_left_map=days_left_map,
        pending=only_pending,
        status_filter=status_filter,
        operator_filter=operator_filter,
    )


@app.route("/calendar", endpoint="calendar_view")
@login_required
def calendar_view():
    days = request.args.get("days", "365").strip()
    try:
        days_int = int(days)
    except ValueError:
        days_int = 365

    today = date.today()
    limit = today + timedelta(days=days_int)

    db = get_db()
    rows = db.execute("SELECT * FROM clients ORDER BY id DESC").fetchall()

    upcoming = []
    for r in rows:
        end_iso = get_end_date_from_client_row(r)
        if not end_iso:
            continue
        end_d = parse_yyyy_mm_dd(end_iso)
        if not end_d:
            continue
        if today <= end_d <= limit:
            upcoming.append((r, (end_d - today).days))

    upcoming.sort(key=lambda t: parse_yyyy_mm_dd(get_end_date_from_client_row(t[0])) or date.max)

    return render_template("calendar.html", upcoming=upcoming, days=days_int, alert_days=ALERT_DAYS)


@app.route("/api/permanencias", endpoint="api_permanencias")
@login_required
def api_permanencias():
    db = get_db()
    rows = db.execute("SELECT * FROM clients ORDER BY id DESC").fetchall()

    out = []
    for r in rows:
        end_iso = get_end_date_from_client_row(r)
        if not end_iso:
            continue

        out.append({
            "id": r["id"],
            "full_name": r["full_name"],
            "phone": r["phone"],
            "email": r["email"],
            "current_operator": r["current_operator"],
            "permanence_end_date": end_iso,
            "days_left": days_until(end_iso),
            "url": url_for("view_client", client_id=r["id"])
        })

    out.sort(key=lambda x: x["permanence_end_date"] or "9999-12-31")
    return jsonify(out)
@app.route("/dashboard")
@login_required
def dashboard():
    db = get_db()

    total_clients = db.execute(
        "SELECT COUNT(*) as total FROM clients"
    ).fetchone()["total"]

    pending_tasks = db.execute("""
        SELECT COUNT(*) as total
        FROM clients
        WHERE pending_tasks IS NOT NULL
        AND TRIM(pending_tasks) != ''
    """).fetchone()["total"]

    month_contracts = db.execute("""
        SELECT COUNT(*) as total
        FROM clients
        WHERE permanence_start_date IS NOT NULL
        AND TRIM(permanence_start_date) != ''
        AND strftime('%Y-%m', permanence_start_date) = strftime('%Y-%m', 'now')
    """).fetchone()["total"]

    upcoming_permanences = 0
    permanence_alerts = []

    rows = db.execute("""
       SELECT id, full_name, permanence_end_date, permanence_end
       FROM clients
    """).fetchall()

    for r in rows:
        end_date = r["permanence_end_date"] or r["permanence_end"]

        if end_date:
            try:
                if "/" in end_date:
                    d, m, y = end_date.split("/")
                    end = date(int(y), int(m), int(d))
                else:
                    y, m, d = end_date.split("-")
                    end = date(int(y), int(m), int(d))
                days_left = (end - date.today()).days

                if 0 <= days_left <= 30:
                    upcoming_permanences += 1
                    permanence_alerts.append({
                        "days_left": days_left,
                        "end_date": end_date,
                        "client_id": r["id"],
                        "full_name": r["full_name"],
                    })
            except:
                pass
    service_counts = db.execute("""
        SELECT service_type, COUNT(*) as total
        FROM clients
        WHERE service_type IS NOT NULL
        AND TRIM(service_type) != ''
        GROUP BY service_type
        ORDER BY total DESC
     """).fetchall()
    recent_clients = db.execute("""
        SELECT id, full_name, current_operator, service_type
        FROM clients
        ORDER BY id DESC
        LIMIT 5
     """).fetchall()
    monthly_contracts_chart = db.execute("""
        SELECT strftime('%Y-%m', permanence_start_date) as month, COUNT(*) as total
        FROM clients
        WHERE permanence_start_date IS NOT NULL
        AND TRIM(permanence_start_date) != ''
        GROUP BY month
        ORDER BY month ASC
    """).fetchall()
    monthly_operator_chart = db.execute("""
        SELECT strftime('%Y-%m', permanence_start_date) as month,
               current_operator,
               COUNT(*) as total
       FROM clients
       WHERE permanence_start_date IS NOT NULL
       AND TRIM(permanence_start_date) != ''
       AND current_operator IS NOT NULL
       AND TRIM(current_operator) != ''
       GROUP BY month, current_operator
       ORDER BY month ASC
   """).fetchall()
    return render_template(
        "dashboard.html",
        total_clients=total_clients,
        pending_tasks=pending_tasks,
        month_contracts=month_contracts,
        upcoming_permanences=len(permanence_alerts),
        permanence_alerts=permanence_alerts,
        service_counts=service_counts,
        recent_clients=recent_clients,
        monthly_contracts_chart=monthly_contracts_chart,
        monthly_operator_chart=monthly_operator_chart,
    )
@app.route("/clients/import", methods=["GET", "POST"])
@login_required
def import_clients():

    if request.method == "POST":

        file = request.files.get("file")

        if not file:
            flash("No se seleccionó archivo")
            return redirect(url_for("import_clients"))

        wb = load_workbook(file)
        ws = wb.active

        db = get_db()

        for row in ws.iter_rows(min_row=2, values_only=True):

            try:
                full_name = row[0]
                dni = row[1]
                phone = row[2]
                operator = row[3]
                status = row[4]
                permanence_end = row[5]
                pending_tasks = row[6]

                db.execute("""
                    INSERT INTO clients (
                        full_name,
                        dni,
                        phone,
                        current_operator,
                        status,
                        permanence_end,
                        permanence_end_date,
                        pending_tasks,
                        service_type,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    full_name,
                    dni,
                    phone,
                    operator,
                    status,
                    permanence_end,
                    permanence_end,
                    pending_tasks,
                    request.form.get("service_type"),
                    datetime.utcnow().isoformat()
                ))

            except Exception as e:
                print("ERROR IMPORTANDO:", e)

        db.commit()

        flash("Clientes importados correctamente")
        return redirect(url_for("clients"))

    return render_template("import_excel.html")

@app.route("/clients/export/excel")
@login_required
def export_clients_excel():
    db = get_db()

    rows = db.execute("SELECT * FROM clients ORDER BY id DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    ws.append([
        "ID", "Nombre", "DNI", "Telefono",
        "Operador", "Fin permanencia", "Estado", "Pendiente"
    ])

    for c in rows:
        end = c["permanence_end_date"] or c["permanence_end"]

        ws.append([
            c["id"],
            c["full_name"],
            c["dni"],
            c["phone"],
            c["current_operator"],
            end,
            c["status"],
            c["pending_tasks"]
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="clientes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/clients/new", methods=["GET", "POST"])
@login_required
def new_client():
    if request.method == "POST":
        db = get_db()

        p_start, p_months, p_end = compute_permanence_end(
            request.form.get("permanence_start_date") or request.form.get("permanence_start"),
            request.form.get("permanence_months"),
            request.form.get("permanence_end_date") or request.form.get("permanence_end"),
        )

        cur = db.execute("""
            INSERT INTO clients (
                full_name, dni, birth_date, phone, address, email,
                current_operator, current_tariff_price,
                permanence,
                permanence_start, permanence_end,
                permanence_start_date, permanence_months, permanence_end_date,
                terminal, sales_done, repairs_done, procedures_done, observations,
                pending_tasks, commercial, status, created_at
            ) VALUES (
                ?, ?, ?, ?, ?, ?,
                ?, ?,
                ?,
                ?, ?,
                ?, ?, ?,
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?
            )
        """, (
            request.form["full_name"],
            request.form["dni"],
            request.form.get("birth_date"),
            request.form.get("phone"),
            request.form.get("address"),
            request.form.get("email"),

            request.form.get("current_operator"),
            request.form.get("current_tariff_price"),

            request.form.get("permanence"),

            p_start, p_end,

            p_start, p_months, p_end,

            request.form.get("terminal"),
            request.form.get("sales_done"),
            request.form.get("repairs_done"),
            request.form.get("procedures_done"),
            request.form.get("observations"),

            request.form.get("pending_tasks"),
            request.form.get("commercial"),
            request.form.get("status"),
            request.form.get("service_type"),
            datetime.utcnow().isoformat()
        ))

        client_id = cur.lastrowid
        db.commit()
        return redirect(url_for("view_client", client_id=client_id))

    return render_template(
        "client_form.html",
        client=None, lines=[], repairs=[], sales=[],
        alert_days=ALERT_DAYS,
        days_left=None
    )


@app.route("/clients/<int:client_id>")
@login_required
def view_client(client_id):
    db = get_db()

    client = db.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    if client is None:
        flash("Cliente no encontrado", "danger")
        return redirect(url_for("clients"))

    lines = db.execute(
        "SELECT * FROM mobile_lines WHERE client_id = ? ORDER BY id DESC",
        (client_id,)
    ).fetchall()

    repairs = db.execute(
        "SELECT * FROM repairs WHERE client_id = ? ORDER BY id DESC",
        (client_id,)
    ).fetchall()

    sales = db.execute(
        "SELECT * FROM sales WHERE client_id = ? ORDER BY id DESC",
        (client_id,)
    ).fetchall()

    end_iso = get_end_date_from_client_row(client)
    du = days_until(end_iso) if end_iso else None

    return render_template(
        "client_form.html",
        client=client,
        lines=lines,
        repairs=repairs,
        sales=sales,
        alert_days=ALERT_DAYS,
        days_left=du
    )


@app.route("/clients/<int:client_id>/update", methods=["POST"])
@login_required
def update_client(client_id):
    db = get_db()

    p_start, p_months, p_end = compute_permanence_end(
        request.form.get("permanence_start_date") or request.form.get("permanence_start"),
        request.form.get("permanence_months"),
        request.form.get("permanence_end_date") or request.form.get("permanence_end"),
    )

    # Si p_end viene vacío, conserva el actual
    if not p_end:
        current = db.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
        old_end = get_end_date_from_client_row(current)
        if old_end:
            p_end = old_end

    db.execute("""
        UPDATE clients SET
            full_name = ?,
            dni = ?,
            birth_date = ?,
            phone = ?,
            address = ?,
            email = ?,
            current_operator = ?,
            current_tariff_price = ?,
            permanence = ?,

            permanence_start = ?,
            permanence_end = ?,

            permanence_start_date = ?,
            permanence_months = ?,
            permanence_end_date = ?,

            terminal = ?,
            sales_done = ?,
            repairs_done = ?,
            procedures_done = ?,
            observations = ?,
            pending_tasks = ?,
            commercial = ?,
            service_type = ?,
            status = ?
        WHERE id = ?
    """, (
        request.form["full_name"],
        request.form["dni"],
        request.form.get("birth_date"),
        request.form.get("phone"),
        request.form.get("address"),
        request.form.get("email"),
        request.form.get("current_operator"),
        request.form.get("current_tariff_price"),
        request.form.get("permanence"),

        p_start,
        p_end,

        p_start,
        p_months,
        p_end,

        request.form.get("terminal"),
        request.form.get("sales_done"),
        request.form.get("repairs_done"),
        request.form.get("procedures_done"),
        request.form.get("observations"),
        request.form.get("pending_tasks"),
        request.form.get("commercial"),
        request.form.get("service_type"),
        request.form.get("status"),
        client_id
    ))

    # -------------------------
    # Guardar líneas móviles SOLO si viene line_count
    # -------------------------
    line_count_raw = request.form.get("line_count", None)

    if line_count_raw is not None:
        db.execute("DELETE FROM mobile_lines WHERE client_id = ?", (client_id,))

        try:
            line_count = int(line_count_raw or 0)
        except ValueError:
            line_count = 0

        for i in range(line_count):
            line_number = (request.form.get(f"line_number_{i}") or "").strip()
            pin = (request.form.get(f"pin_{i}") or "").strip()
            puk = (request.form.get(f"puk_{i}") or "").strip()
            icc = (request.form.get(f"icc_{i}") or "").strip()
            account = (request.form.get(f"account_{i}") or "").strip()
            line_perm_end = (request.form.get(f"line_perm_end_{i}") or "").strip()

            if not (line_number or pin or puk or icc or account or line_perm_end):
                continue

            db.execute("""
                INSERT INTO mobile_lines (
                    client_id, line_number, pin, puk, icc,
                    google_or_iphone_account, permanence_end_date, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                client_id,
                line_number,
                pin,
                puk,
                icc,
                account,
                line_perm_end,
                datetime.utcnow().isoformat()
            ))

    db.commit()
    flash("Cliente actualizado", "success")
    return redirect(url_for("view_client", client_id=client_id))


@app.route("/clients/<int:client_id>/repairs/add", methods=["POST"])
@login_required
def add_repair(client_id):
    db = get_db()

    date_ = request.form.get("repair_date")
    model = request.form.get("repair_model")
    repair = request.form.get("repair_text")

    cost_raw = (request.form.get("repair_cost") or "").strip()
    cost = None
    if cost_raw:
        try:
            cost = float(cost_raw.replace(",", "."))
        except ValueError:
            cost = None

    db.execute("""
        INSERT INTO repairs (client_id, date, model, repair, cost, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        client_id,
        date_,
        model,
        repair,
        cost,
        datetime.utcnow().isoformat()
    ))
    db.commit()
    return redirect(url_for("view_client", client_id=client_id))


@app.route("/clients/<int:client_id>/repairs/<int:repair_id>/delete", methods=["POST"])
@login_required
def delete_repair(client_id, repair_id):
    db = get_db()
    db.execute("DELETE FROM repairs WHERE id = ? AND client_id = ?", (repair_id, client_id))
    db.commit()
    return redirect(url_for("view_client", client_id=client_id))


@app.route("/clients/<int:client_id>/sales/add", methods=["POST"])
@login_required
def add_sale(client_id):
    db = get_db()

    date_ = request.form.get("sale_date")
    item = request.form.get("sale_item")
    operator = request.form.get("sale_operator")

    amount_raw = (request.form.get("sale_amount") or "").strip()
    amount = None
    if amount_raw:
        try:
            amount = float(amount_raw.replace(",", "."))
        except ValueError:
            amount = None

    notes = request.form.get("sale_notes")

    db.execute("""
        INSERT INTO sales (client_id, date, item, operator, amount, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        client_id,
        date_,
        item,
        operator,
        amount,
        notes,
        datetime.utcnow().isoformat()
    ))
    db.commit()
    return redirect(url_for("view_client", client_id=client_id))


@app.route("/clients/<int:client_id>/sales/<int:sale_id>/delete", methods=["POST"])
@login_required
def delete_sale(client_id, sale_id):
    db = get_db()
    db.execute("DELETE FROM sales WHERE id = ? AND client_id = ?", (sale_id, client_id))
    db.commit()
    return redirect(url_for("view_client", client_id=client_id))


@app.route("/clients/<int:client_id>/delete", methods=["POST"])
@login_required
def delete_client(client_id):
    db = get_db()
    db.execute("DELETE FROM clients WHERE id = ?", (client_id,))
    db.commit()
    flash("Cliente eliminado", "success")
    return redirect(url_for("clients"))


if __name__ == "__main__":
    app.run(debug=True)
